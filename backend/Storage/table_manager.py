import json
import os
import csv
from openpyxl import load_workbook
import bisect
import pickle
import copy
from functools import cmp_to_key

class TableManager:
    def __init__(self,db_path="./Data",pickle_path="./Pickles"):

      if not os.path.isabs(db_path):
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        db_path = os.path.join(base_dir, db_path.lstrip('./'))
    
      if not os.path.isabs(pickle_path):
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        pickle_path = os.path.join(base_dir, pickle_path.lstrip('./'))
    
   
      self.db_path = db_path
      self.pickle_path = pickle_path
      self.memory_indexes = {}
    
   
      if not os.path.exists(db_path):
        os.makedirs(db_path)
        print(f"folder {db_path} ready 👍")
    
      if not os.path.exists(pickle_path):
        os.makedirs(pickle_path)
        print(f"folder {pickle_path} ready 👍")
    
      self.column_order = {}
      
    def create_table(self,table_name,columns,index_hints=None):
        table_file=f"{self.db_path}/{table_name}.json"

        if os.path.exists(table_file):
             print(f"{table_name} file is already exists")
             return False
        
            
        hashing_index=[]
        b_tree_index=[]
        unique_patterns = {
        "id", "email", "username", "phone", "mobile",
        "userid", "user_id", "employee_id", "customer_id",
        "ssn", "passport", "license_number",
        "uuid", "guid", "token", "api_key"
    }
         
        duplicate_patterns = {
        "name", "first_name", "last_name", "middle_name",
        "age", "salary", "price", "amount", "quantity",
        "city", "state", "country", "address", "zip",
        "department", "role", "status", "category", "type",
        "date", "created_at", "updated_at", "timestamp"
    }


        for col_name,col_type in columns.items():
             if index_hints and col_name in index_hints:
                  hint = index_hints[col_name]
                  if hint == "HASH":
                     hashing_index.append(col_name)
                     print(f"   🚀 {col_name} → Hash (user specified)")
                     continue
                  elif hint == "BTREE":
                       b_tree_index.append(col_name)
                       print(f"   🌳 {col_name} → B-tree (user specified)")
                       continue
                  elif hint == "NONE":
                       print(f"   ⚪ {col_name} → No index (user specified)")
                       continue
             if col_type in ["INT", "FLOAT"] or col_name.lower() in unique_patterns:
                 hashing_index.append(col_name)
             elif col_name.lower()  in duplicate_patterns:
                  b_tree_index.append(col_name)
             else:
                 b_tree_index.append(col_name)
                 print(f"   🌳 {col_name} → B-tree (default)")
        
       

         schema={"columns":columns,
                "indexes":{
                   "hashing": hashing_index,
                   "b_tree": b_tree_index
                },
                "rows": []
                }
         
        
         with open(table_file,"w",encoding="utf-8") as f:
            json.dump(schema,f,indent=2,ensure_ascii=False)
          
         self._load_indexes(table_name)
         print(f"File Ready {table_name} 👍")
         
       
        
    def _load_indexes(self,table_name):
       table_file = f"{self.db_path}/{table_name}.json"
       index_file=f"{self.pickle_path}/{table_name}_indexes.pkl"
       with open(table_file, "r", encoding="utf-8") as f:
        table_data = json.load(f)
       if os.path.exists(index_file):
        with open (index_file,"rb") as f:
           self.memory_indexes[table_name]=pickle.load(f)
           return
       self.memory_indexes[table_name]={
           "hash":{},
            "b_tree":{}
       }
       for col in table_data["indexes"]["hashing"]:
            self.memory_indexes[table_name]["hash"][col]={}
       for col in table_data["indexes"]["b_tree"]:
          self.memory_indexes[table_name]["b_tree"][col]={
             "keys":[],
             "values":{}

          }
       self._save_indexes_to_disk(table_name)
       
    def _save_indexes_to_disk(self,table_name):
       index_file=f"{self.pickle_path}/{table_name}_indexes.pkl"
       with open (index_file,"wb") as f:
          pickle.dump(self.memory_indexes[table_name],f)       

    def load_pickle(self,table):
       pickle_file = f"{self.pickle_path}/{table}_indexes.pkl"
       if not os.path.exists(pickle_file):
        print(f"❌ Pickle file not found for table '{table}'")
        return None
    
       with open(pickle_file, "rb") as f:
        indexes = pickle.load(f)
    
        print(f"\n{'='*60}")
       print(f"📦 PICKLE FILE: {table}_indexes.pkl")
       print(f"{'='*60}")
    
       hash_indexes = indexes.get("hash", {})
       print(f"\n🚀 Hash Indexes ({len(hash_indexes)}):")
       for col, index in hash_indexes.items():
        print(f"   {col}: {len(index)} entries")
        if index:
            sample = list(index.items())[:3]
            for key, value in sample:
                print(f"      {key} → {value}")
            if len(index) > 3:
                print(f"      ... ({len(index) - 3} more)")
      
        btree_indexes = indexes.get("b_tree", {})
       print(f"\n🌳 B-Tree Indexes ({len(btree_indexes)}):")
       for col, index in btree_indexes.items():
        keys = index.get("keys", [])
        values = index.get("values", {})  # ← GET VALUES!
        print(f"   {col}: {len(keys)} keys")
        if keys:
            print(f"      Keys: {keys[:5]}{'...' if len(keys) > 5 else ''}")
            for key in keys[:3]:
                val = values.get(key, [])
                print(f"      {key} → {val}")
            if len(keys) > 3:
                print(f"      ... ({len(keys) - 3} more)")
    
       print(f"\n{'='*60}\n")
    
       result = {
        "table": table,
        "hash_indexes": [],
        "btree_indexes": []
    }
    
    # Hash indexes
       for col, index in hash_indexes.items():
        entries = []
        for key, value in list(index.items())[:10]:
            entries.append({"key": str(key), "value": str(value)})
        
        result["hash_indexes"].append({
            "column": col,
            "total": len(index),
            "entries": entries
        })
    
    # B-tree indexes
       for col, index in btree_indexes.items():
        keys = index.get("keys", [])
        values = index.get("values", {}) 
        
        entries = []
        for key in keys[:10]:
            val = values.get(key, []) 
            entries.append({"key": str(key), "value": str(val)})
        
        result["btree_indexes"].append({
            "column": col,
            "total": len(keys),
            "entries": entries
        })
    
       return result  
        
          
        
    def _add_to_index(self,table_name,values,save_to_disk=True):
       table_file = f"{self.db_path}/{table_name}.json"
       with open(table_file, "r", encoding="utf-8") as f:
         table_data = json.load(f)
       columns=table_data.get("columns")
       all_rows=[]
       for row in values:
        row_dict=dict(zip(columns.keys(),row))
        all_rows.append(row_dict)
       for row in all_rows:
          for col in self.memory_indexes[table_name]["hash"]: 
             value=row.get(col)
             self.memory_indexes[table_name]["hash"][col][value]=row
          for col in self.memory_indexes[table_name]["b_tree"]:
             value=row.get(col)
             b_tree=self.memory_indexes[table_name]["b_tree"][col]
             if value not in b_tree["values"]:
                bisect.insort(b_tree["keys"],value)
                b_tree["values"][value]=[]
             row_exists = False
             for existing_row in b_tree["values"][value]:
                 if existing_row == row:
                     row_exists = True
                     break
             if not row_exists:
                  b_tree["values"][value].append(row)
       if save_to_disk:
          self._save_indexes_to_disk(table_name)

    def insert_rows(self, table_name, values):
    
    
     table_file = f"{self.db_path}/{table_name}.json"
    
     if not os.path.exists(table_file):
        print(f"❌ {table_name} does not exist")
        return
    
     with open(table_file, "r", encoding="utf-8") as f:
        table_data = json.load(f)
    
     columns = table_data["columns"]
     columns_name = list(columns.keys())
     for val in values:
       if len(columns_name) != len(val):
         print(f"❌ Column count mismatch: expected {len(columns_name)}, got {len(val)}")
         return

     if len(values) > 0 and isinstance(values[0], list):
        rows_to_insert = values
     else:
        rows_to_insert = [values]
        
     if table_name not in self.memory_indexes:
            self._load_indexes(table_name)
     
     for row_values in rows_to_insert:
        converted_values = []
        
        for i, (col_name,col_type) in enumerate(columns.items()):
            value = row_values[i]
            
            if col_type =="INT":
                converted_values.append(int(value))
            elif col_type =="FLOAT":
                converted_values.append(float(value))
            else:
                converted_values.append(str(value))
        
        self._add_to_index(table_name,[converted_values],save_to_disk=True)
        table_data["rows"].append(converted_values)
    
     with open(table_file, "w", encoding="utf-8") as f:
        json.dump(table_data, f, indent=2, ensure_ascii=False)
    
     print(f"✅ {len(rows_to_insert)} Rows successfully added")
    def insert_col(self,table_name,values):
       table_file = f"{self.db_path}/{table_name}.json"
       if not os.path.exists(table_file):
          print(f"❌ {table_name} does not exist")
          return
       with open(table_file, "r", encoding="utf-8") as f:
            table_data = json.load(f)
            columns=table_data["columns"]
       if len(columns)!=0:
          print("❌ columns are not empty")
          return
       schema={"columns":values,
                "rows": []
                }
       with open(table_file,"w",encoding="utf-8") as f:
            json.dump(schema,f,indent=2,ensure_ascii=False)
       print(f"✅ columns are inserted to {table_name} table")
    
    def addmore_col(self,table_name,values):
         table_file = f"{self.db_path}/{table_name}.json"
         if not os.path.exists(table_file):
            print(f"❌ {table_name} does not exist")
            return
         unique_patterns = {
        "id", "email", "username", "phone", "mobile",
        "userid", "user_id", "employee_id", "customer_id",
        "ssn", "passport", "license_number",
        "uuid", "guid", "token", "api_key"
    }
         with open(table_file, "r", encoding="utf-8") as f:
               table_data = json.load(f)
               columns=table_data["columns"]
               rows=table_data["rows"]
               indexes = table_data.get("indexes", {"hashing": [], "b_tree": []})
         new_columns = {}     

         for col_name, col_type in values.items():
           if col_name in columns:
            print(f"❌ Column '{col_name}' already exists in {table_name}")
            continue  
           else:
            new_columns[col_name] = col_type
     
         if not new_columns:
           print(f"⚠️ No new columns to add")
           return 

         for col_name,col_type in new_columns.items():
            columns[col_name]=col_type
            table_data["columns"] = columns
            columns_name = list(columns.keys())
            all_rows=self.get_allrows(table_name)
            upd_rows=[]
            for row in all_rows:
               for new_col in new_columns.keys():
                  row[new_col]=None
               new_row=[row[col] for col in columns_name]
               upd_rows.append(new_row)
         table_data["rows"]=upd_rows

         for col_name in new_columns.keys():
            if col_name.lower() in unique_patterns:
               if col_name not in indexes["hashing"]:  
                indexes["hashing"].append(col_name)
            else:
             if col_name not in indexes["b_tree"]: 
                indexes["b_tree"].append(col_name)
    
         table_data["indexes"] = indexes
         with open(table_file,"w",encoding="utf-8") as f:
            json.dump(table_data,f,indent=2,ensure_ascii=False)
         
         if table_name not in self.memory_indexes:
            self._load_indexes(table_name)
         for col_name,col_type in new_columns.items():
            if col_name.lower() in unique_patterns:
               self.memory_indexes[table_name]["hash"][col_name] = {}
            else:
               self.memory_indexes[table_name]["b_tree"][col_name] = {
                "keys": [],
                "values": {}
            }
               
         self._save_indexes_to_disk(table_name)
         print(f"✅ {len(new_columns)} column(s) are added to {table_name} table")         
    
    def update_col(self, table_name, col, col_type):
   
       table_file = f"{self.db_path}/{table_name}.json"
       pickle_file = f"{self.pickle_path}/{table_name}_indexes.pkl"
    
       if not os.path.exists(table_file):
        print(f"❌ {table_name} does not exist")
        return
    
   
       with open(table_file, "r", encoding="utf-8") as f:
        table_data = json.load(f)
        columns = table_data["columns"]
        rows = table_data["rows"]
        indexes = table_data.get("indexes", {"hashing": [], "b_tree": []})
    
    
       if col not in columns: 
        print(f"❌ column {col} does not exist")
        return
    
    
       old_type = columns[col]
       new_type = col_type.upper()
    
       if old_type == new_type:
        print(f"⚠️ Column already has type '{new_type}'")
        return
    
   
       columns[col] = new_type
       table_data["columns"] = columns
    
    
       if rows:
        columns_name = list(columns.keys())
        col_indx = columns_name.index(col)
        conv_count = 0
        for i, row in enumerate(rows):
            if col_indx >= len(row):
             continue
            old_val = row[col_indx]
            new_val = self._convert_val(old_val, new_type)
            row[col_indx] = new_val
            conv_count += 1
        table_data["rows"] = rows
        
        
   
       was_hash = col in indexes.get("hashing", [])
       was_btree = col in indexes.get("b_tree", [])
    
       unique_patterns = {
        "id", "email", "username", "phone", "mobile",
        "userid", "user_id", "employee_id", "customer_id",
        "ssn", "passport", "license_number",
        "uuid", "guid", "token", "api_key"
    }
    
       duplicate_patterns = {
        "name", "first_name", "last_name", "middle_name",
        "age", "salary", "price", "amount", "quantity",
        "city", "state", "country", "address", "zip",
        "department", "role", "status", "category", "type",
        "date", "created_at", "updated_at", "timestamp"
    }
    
       should_be_hash = col.lower() in unique_patterns
       should_be_btree = (col.lower() in duplicate_patterns or 
          new_type in ["int", "float", "str", "text"]) 
    
    
       if should_be_hash and not was_hash:
        if was_btree:
            indexes["b_tree"].remove(col)
        if col not in indexes["hashing"]:
            indexes["hashing"].append(col)
        
    
    
       elif should_be_btree and not was_btree:
        if was_hash:
            indexes["hashing"].remove(col)
        if col not in indexes["b_tree"]:
            indexes["b_tree"].append(col)
    
       table_data["indexes"] = indexes
    
       with open(table_file, "w", encoding="utf-8") as f:
        json.dump(table_data, f, indent=2, ensure_ascii=False)
    
       print(f"✅ column {col} type is updated to {col_type} in {table_name} table")
    
       if table_name not in self.memory_indexes:
        self._load_indexes(table_name)
    
       if was_hash and col in self.memory_indexes[table_name]["hash"]:
          del self.memory_indexes[table_name]["hash"][col]
       if was_btree and col in self.memory_indexes[table_name]["b_tree"]:
          del self.memory_indexes[table_name]["b_tree"][col]
       if should_be_hash:
          self._create_hash_index(table_name, col)
       elif should_be_btree:
          self._create_btree_index(table_name, col)
    
       self._save_indexes_to_disk(table_name)
       print(f"   💾 Memory indexes saved")
         
    def _convert_val(self,value,target_type):
       if value is None:
          return None
       target_type=target_type.upper()
       if target_type=="INT":
          if isinstance(value,int):
             return value
          elif isinstance(value,float):
             return int(value)
          elif isinstance(value,str):
             return int(float(value))
       elif target_type=="FLOAT":
          if isinstance(value,float):
             return value
          elif isinstance(value,int):
             return float(value)
          elif isinstance(value,str):
             return float(value)
       elif target_type in ["STR","TEXT"]:
          return str(value)
       else:
          return value
    def _create_hash_index(self, table_name, col):
       all_rows = self.get_allrows(table_name)
       hash_index = {}
    
       for row in all_rows:
          value = row.get(col)
        
          if value is not None:
            hash_index[value] = row
       self.memory_indexes[table_name]["hash"][col] = hash_index
    
    def _create_btree_index(self, table_name, col):
    
       all_rows = self.get_allrows(table_name)
       btree_index = {"keys": [], "values": {}}
       values_list = []
       for row in all_rows:
           value = row.get(col)
           if value is not None:
              values_list.append((value, row))
    
       values_list.sort(key=lambda x: x[0])
    
       for value, row in values_list:
        if value not in btree_index["values"]:
            btree_index["keys"].append(value)
            btree_index["values"][value] = []
        btree_index["values"][value].append(row)
    
       self.memory_indexes[table_name]["b_tree"][col] = btree_index

    def insert_rows_csv(self, table_name, values):
    
    
      table_file = f"{self.db_path}/{table_name}.json"
    
      if not os.path.exists(table_file):
        print(f"❌ {table_name} does not exist")
        return
    
      with open(table_file, "r", encoding="utf-8") as f:
        table_data = json.load(f)
    
      columns = table_data["columns"]
      columns_name = list(columns.keys())
      for val in values:
       if len(columns_name) != len(val):
         return

      if len(values) > 0 and isinstance(values[0], list):
        rows_to_insert = values
      else:
        rows_to_insert = [values]

      if table_name not in self.memory_indexes:
            self._load_indexes(table_name)
    
    
      for row_values in rows_to_insert:
        converted_values = []
        
        for i, (col_name,col_type) in enumerate(columns.items()):
            value = row_values[i]
            
            if col_type == "INT":
                converted_values.append(int(value))
            elif col_type == "FLOAT":
                converted_values.append(float(value))
            else:
                converted_values.append(str(value))

          
        self._add_to_index(table_name,[converted_values],save_to_disk=True)
        table_data["rows"].append(converted_values)
    
      with open(table_file, "w", encoding="utf-8") as f:
        json.dump(table_data, f, indent=2, ensure_ascii=False)
    
      return True
   
    def a(self,table_name,col):
       print(f"table_name{table_name},{col}")

    def delete(self,table_name):
       table_file=f"{self.db_path}/{table_name}.json"
       table_pickle=f"{self.pickle_path}/{table_name}_indexes.pkl"

       if not os.path.exists(table_file):
          print("file not exits ❌")
          return
       else:
          os.remove(table_file)
          os.remove(table_pickle)
          print(f"✅ {table_name } file is deleted")

    def get_allrows(self,table_name,col=None):
       table_file=f"{self.db_path}/{table_name}.json"
       if not os.path.exists(table_file):
          print("files not exists ❌")
          return[]
       with open (table_file,"r",encoding="utf-8") as f:
          table_data=json.load(f)
          columns=table_data.get("columns",{})
          rows=table_data.get("rows",[])
          if len(rows)==0:
             print("rows are empty")
             return []
          columns_name=list(columns.keys())
          if col is not None and col not in columns_name:
             print("column not found")
             return []
          all_rows=[]
          for row in rows:
             row_dict={}
             for i, col_name in enumerate(columns_name):
              row_dict[col_name] = row[i]
             all_rows.append(row_dict)
          
          return all_rows
       
    def row_equal(self,row1,row2):
       if set(row1.keys()) != set(row2.keys()):
        return False
    
    
       for key in row1.keys():
        if row1[key] != row2[key]:
            return False
    
       return True

       
    def update_indexes(self, table_name, old_rows, new_rows, upd_col):
       
       hash_indexes = self.memory_indexes[table_name]["hash"]
       btree_indexes = self.memory_indexes[table_name]["b_tree"]
    
       for old_row, new_row in zip(old_rows, new_rows):
        
        
        if upd_col in hash_indexes:
            old_value = old_row[upd_col]
            new_value = new_row[upd_col]
            
           
            if old_value in hash_indexes[upd_col]:
                del hash_indexes[upd_col][old_value]
            
            
            hash_indexes[upd_col][new_value] = new_row
        
        
        if upd_col in btree_indexes:
            old_value = old_row[upd_col]
            new_value = new_row[upd_col]
            
            
            if old_value in btree_indexes[upd_col]["values"]:
                value_list = btree_indexes[upd_col]["values"][old_value]

                for i in range(len(value_list) - 1, -1, -1):
                    if self.row_equal(value_list[i], old_row):
                        value_list.pop(i)
                        break
                
               
                if not value_list:
                    del btree_indexes[upd_col]["values"][old_value]
                    btree_indexes[upd_col]["keys"].remove(old_value)
            
            
            if new_value not in btree_indexes[upd_col]["values"]:
                bisect.insort(btree_indexes[upd_col]["keys"], new_value)
                btree_indexes[upd_col]["values"][new_value] = []
            
            btree_indexes[upd_col]["values"][new_value].append(new_row)
        
        
        for col in hash_indexes:
            if col != upd_col:
                for key in list(hash_indexes[col].keys()):

                    if self.row_equal(hash_indexes[col][key], old_row):
                        hash_indexes[col][key] = new_row
            
        
        for col in btree_indexes:
            if col != upd_col:
                for value_list in btree_indexes[col]["values"].values():
                    for i, row in enumerate(value_list):
                        
                        if self.row_equal(row, old_row):
                            value_list[i] = new_row
       
    def update_rows(self, table_name, upd_col, upd_value, where_col, op, where_value):
       
       table_file = f"{self.db_path}/{table_name}.json"
    
       if not os.path.exists(table_file):
         print(f"❌ {table_name} does not exist")
         return
    
       with open(table_file, "r", encoding="utf-8") as f:
        table_data = json.load(f)
    
       columns = table_data.get("columns", {})
       columns_name = list(columns.keys())
    
       if upd_col not in columns_name:
        print(f"❌ Column {upd_col} does not exist")
        return
    
       if where_col not in columns_name:
        print(f"❌ Column {where_col} does not exist")
        return
    
       if table_name not in self.memory_indexes:
        self._load_indexes(table_name)
    
    
       where_col_type = columns.get(where_col)
    
       if where_col_type == "INT":
         converted_where_value = int(where_value)
         
       elif where_col_type == "FLOAT":
          converted_where_value = float(where_value)
       else:
           converted_where_value = str(where_value)
         
    
    
       matching_rows = []
    
       if op == "=" and where_col in self.memory_indexes[table_name]["hash"]:
        print(f"🚀 Using Hash Index on '{where_col}'")
        result = self.memory_indexes[table_name]["hash"][where_col].get(converted_where_value)
        if result:

            matching_rows = [copy.deepcopy(result)]
    
       elif op in [">", ">=", "<", "<=", "="] and where_col in self.memory_indexes[table_name]["b_tree"]:
        print(f"🌳 Using B-Tree Index on '{where_col}'")
        btree = self.memory_indexes[table_name]["b_tree"][where_col]
        
        if op == "=":
            if converted_where_value in btree["values"]:
                
                matching_rows = [copy.deepcopy(row) for row in btree["values"][converted_where_value]]
        elif op == ">":
            start_idx = bisect.bisect_right(btree["keys"], converted_where_value)
            for i in range(start_idx, len(btree["keys"])):
                matching_rows.extend([copy.deepcopy(row) for row in btree["values"][btree["keys"][i]]])
        elif op == ">=":
            start_idx = bisect.bisect_left(btree["keys"], converted_where_value)
            for i in range(start_idx, len(btree["keys"])):
                matching_rows.extend([copy.deepcopy(row) for row in btree["values"][btree["keys"][i]]])
        elif op == "<":
            end_idx = bisect.bisect_left(btree["keys"], converted_where_value)
            for i in range(0, end_idx):
                matching_rows.extend([copy.deepcopy(row) for row in btree["values"][btree["keys"][i]]])
        elif op == "<=":
            end_idx = bisect.bisect_right(btree["keys"], converted_where_value)
            for i in range(0, end_idx):
                matching_rows.extend([copy.deepcopy(row) for row in btree["values"][btree["keys"][i]]])
    
       elif op == "!=" and where_col in self.memory_indexes[table_name]["hash"]:
        print(f"🚀 Using Hash Index (inverted)")
        for key, row in self.memory_indexes[table_name]["hash"][where_col].items():
            if key != converted_where_value:
                
                matching_rows.append(copy.deepcopy(row))
    
       else:
        print(f"⚠️ Using linear search")
        all_rows_temp = self.get_allrows(table_name)
        if all_rows_temp is None:
            return
        
        for row_dict in all_rows_temp:
            where_row_value = row_dict.get(where_col)
            
            if where_col_type == "INT":
                where_row_value = int(where_row_value)
            elif where_col_type == "FLOAT":
                where_row_value = float(where_row_value)
            else:
                where_row_value = str(where_row_value)
            
            condition_met = False
            
            if op == "=" and where_row_value == converted_where_value:
                condition_met = True
            elif op == ">" and where_row_value > converted_where_value:
                condition_met = True
            elif op == "<" and where_row_value < converted_where_value:
                condition_met = True
            elif op == ">=" and where_row_value >= converted_where_value:
                condition_met = True
            elif op == "<=" and where_row_value <= converted_where_value:
                condition_met = True
            elif op == "!=" and where_row_value != converted_where_value:
                condition_met = True
            
            if condition_met:
                matching_rows.append(row_dict)
    
       if len(matching_rows) == 0:
        print(f"⚠️ No rows match WHERE {where_col} {op} {where_value}")
        return 0
    

       old_matching_rows = [row.copy() for row in matching_rows]
    
   
       upd_col_type = columns.get(upd_col)
    
       if upd_col_type == "INT":
        new_value = int(upd_value)
       elif upd_col_type == "FLOAT":
        new_value = float(upd_value)
       else:
        new_value = str(upd_value)
    
    
       upd_count = 0
       updated_info = []
    
       for row_dict in matching_rows:
        old_value = row_dict.get(upd_col)
        row_dict[upd_col] = new_value
        upd_count += 1
        updated_info.append({"old": old_value, "new": new_value})
    
   
       all_rows = self.get_allrows(table_name)
    
    
       id_col = columns_name[0]  
       rows_to_update = {row[id_col] for row in matching_rows}
    
       for row_dict in all_rows:
        if row_dict[id_col] in rows_to_update:
            row_dict[upd_col] = new_value
    
    
       new_rows = []
       for row_dict in all_rows:
        row_list = [row_dict.get(col) for col in columns_name]
        new_rows.append(row_list)
    
       table_data["rows"] = new_rows
    
    
       with open(table_file, "w", encoding="utf-8") as f:
        json.dump(table_data, f, indent=2, ensure_ascii=False)
    
    
       print(f"🔄 Updating indexes...")
       self.update_indexes(table_name, old_matching_rows, matching_rows, upd_col)
       self._save_indexes_to_disk(table_name)
    
       print(f"✅ {upd_count} row(s) updated in '{table_name}'")
       if updated_info:
        first = updated_info[0]
        print(f"   {upd_col}: {first['old']} → {first['new']}")
      
       return upd_count
       
       
    def read_csv(self,table_name,csv_file):
       table_file = f"{self.db_path}/{table_name}.json"
    
       if not os.path.exists(table_file):
        print(f"❌ {table_name} does not exist")
        return
      
       
       with open (csv_file,"r",encoding="utf=8") as f:
          csv_reader=csv.reader(f)
          headers=next(csv_reader)
          headers=[h.strip()for h in headers]
          print(f"📋 Columns found: {headers}")
          rows=list(csv_reader)
          rows=[[value.strip()for value in row]for row in rows]
          print(f"📊 Total rows: {len(rows)}")

       if len(rows) == 0:
            print("⚠️ No data rows found")
            return
       inserted=False
       inserted_count=0
       for row in (rows):
          insert_row=self.insert_rows_csv(table_name,[row])
          if insert_row==True:
            inserted=True 
            inserted_count+=1
          else:
             break
       
       if inserted==True:
          print(f"✅ total {inserted_count} rows are inserted to {table_name} table")
       elif inserted==False:
          print(f"❌ Column mismatch btw table and csv file")
          
       return inserted_count
    
    def insert_from_excel(self,table_name,excel_file,sheet_name=None):
       table_file = f"{self.db_path}/{table_name}.json"
    
       if not os.path.exists(table_file):
        print(f"❌ {table_name} does not exist")
        return
       
       workbook=load_workbook(excel_file)
       available_sheets = workbook.sheetnames
       print(f"📋 Available sheets: {', '.join(available_sheets)}")
       if sheet_name:
            if sheet_name in available_sheets:
                sheet = workbook[sheet_name]
                print(f"✅ Using sheet: '{sheet_name}'")
            else:
                print(f"⚠️ Sheet '{sheet_name}' not found!")
                print(f"🔄 Using first sheet: '{available_sheets[0]}'")
                sheet = workbook[available_sheets[0]]
       else:
          sheet=workbook.active
       rows=list(sheet.iter_rows(values_only=True))
       if len(rows)<2:
          print("⚠️ No data found")
          return
       headers=rows[0]
       print(f"📋 Columns found: {headers}")
       data_rows=rows[1:]

       inserted_count=0
       inserted=False
       for row in data_rows:
          row_list=list(row)
          insert_row=self.insert_rows_csv(table_name,[row_list])
          if insert_row==True:
             inserted=True
             inserted_count+=1
          else:
             break
       if inserted==True:
          print(f"✅ total {inserted_count} rows are inserted to {table_name} table")
       elif inserted==False:
          print(f"❌ Column mismatch btw table and excel file")
    
    def delete_all_rows(self,table_name):
       table_file=f"{self.db_path}/{table_name}.json"
       with open(table_file,"r",encoding="utf-8") as f:
          table_data=json.load(f)
          rows=table_data["rows"]
          if len(rows)==0:
            print("❌ rows are already empty")
            return
          rows.clear()
          table_data["rows"]=rows
       with open(table_file,"w",encoding="utf-8") as f:
           json.dump(table_data,f,indent=2,ensure_ascii=False)
       print(f"✅ all rows are deleted of table {table_name} ")
    def delete_all_col(self,table_name):
         table_file=f"{self.db_path}/{table_name}.json"
         with open(table_file,"r",encoding="utf-8") as f:
            table_data=json.load(f)
            columns=table_data["columns"]
            if len(columns)==0:
               print("❌ columns are already empty")
               return
            columns.clear()
            table_data["columns"]=columns
         with open(table_file,"w",encoding="utf-8") as f:
            json.dump(table_data,f,indent=2,ensure_ascii=False)
         print(f"✅ all columns are deleted of table {table_name} ")

    def delete_row(self,table_name,col,op,value):  
       table_file=f"{self.db_path}/{table_name}.json"
       with open(table_file, "r", encoding="utf-8") as f:
         table_data = json.load(f)
         columns = table_data.get("columns", {})
         columns_name = list(columns.keys())
         if col not in columns_name:
            print(f"{col}. not in columns")
         if table_name not in self.memory_indexes:
            self._load_indexes(table_name)

         where_col_type = columns.get(col)
    
       if where_col_type == "INT":
         converted_where_value = int(value)
       elif where_col_type == "FLOAT":
          converted_where_value = float(value)
       else:
           converted_where_value = str(value)

       row_to_delete=[]       

       if op == "=" and col in self.memory_indexes[table_name]["hash"]:
        print(f"🚀 Using Hash Index on '{col}'")
        result = self.memory_indexes[table_name]["hash"][col].get(converted_where_value)
        if result:
            row_to_delete = [copy.deepcopy(result)]
            

    
       elif op in [">", ">=", "<", "<=", "="] and col in self.memory_indexes[table_name]["b_tree"]:
        print(f"🌳 Using B-Tree Index on '{col}'")
        btree = self.memory_indexes[table_name]["b_tree"][col]
        
        if op == "=":
            if converted_where_value in btree["values"]:
                
                row_to_delete = [copy.deepcopy(row) for row in btree["values"][converted_where_value]]
                
        elif op == ">":
            start_idx = bisect.bisect_right(btree["keys"], converted_where_value)
            for i in range(start_idx, len(btree["keys"])):
                row_to_delete.extend([copy.deepcopy(row) for row in btree["values"][btree["keys"][i]]])
                
   
        elif op == ">=":
            start_idx = bisect.bisect_left(btree["keys"], converted_where_value)
            for i in range(start_idx, len(btree["keys"])):
                row_to_delete.extend([copy.deepcopy(row) for row in btree["values"][btree["keys"][i]]])
                
        elif op == "<":
            end_idx = bisect.bisect_left(btree["keys"], converted_where_value)
            for i in range(0, end_idx):
                row_to_delete.extend([copy.deepcopy(row) for row in btree["values"][btree["keys"][i]]])
                
        elif op == "<=":
            end_idx = bisect.bisect_right(btree["keys"], converted_where_value)
            for i in range(0, end_idx):
                row_to_delete.extend([copy.deepcopy(row) for row in btree["values"][btree["keys"][i]]])
               
    
       elif op == "!=" and col in self.memory_indexes[table_name]["hash"]:
        print(f"🚀 Using Hash Index (inverted)")
        for key, row in self.memory_indexes[table_name]["hash"][col].items():
            if key != converted_where_value:
                
                row_to_delete.append(copy.deepcopy(row))
                
       else:
          print(f"⚠️ Using linear search")
          all_rows=self.get_allrows(table_name)
          for row in all_rows:
             if op=="=" and str(row.get(col))==str(value):
               row_to_delete.append(row)
               
             elif op==">" and float(row.get(col))>float(value):
                row_to_delete.append(row)
                
             elif op=="<" and float(row.get(col))<float(value):
               row_to_delete.append(row)
               
             elif op==">=" and float(row.get(col))>=float(value):
              row_to_delete.append(row)
              
             elif op=="<=" and float(row.get(col))<=float(value):
               row_to_delete.append(row)
               
             elif op=="!=" and str(row.get(col))!=str(value):
              row_to_delete.append(row) 
              

       if len(row_to_delete) == 0:
        print(f"⚠️ No rows match WHERE {col} {op} {value}")
        return 0
       all_rows = self.get_allrows(table_name)
       id_col = columns_name[0]
       ids_to_delete = {row[id_col] for row in row_to_delete}
       rows_to_keep = [row for row in all_rows if row[id_col] not in ids_to_delete]
       
       new_rows=[]
       for row_dict in rows_to_keep:
          row_list = [row_dict.get(col_name) for col_name in columns_name]
          new_rows.append(row_list)
         
       table_data["rows"]=new_rows
       with open(table_file,"w",encoding="utf-8")as f:
            json.dump(table_data,f,indent=2,ensure_ascii=False)
       print(f"✅ deleted {len(row_to_delete)} rows")
       print(f"👍there are {len(new_rows)} rows remaining in the table")
       return rows_to_keep

    def delete_from_indexes(self,table_name,rows_to_delete):
       hash_indexes = self.memory_indexes[table_name]["hash"]
       btree_indexes = self.memory_indexes[table_name]["b_tree"]
    
       deleted_count = 0
    
       for row in rows_to_delete:
          for col in hash_indexes:
            col_value = row.get(col)
         
            
            if col_value in hash_indexes[col]:
               
                stored_row = hash_indexes[col][col_value]
                
                if self._rows_equal(stored_row, row):
                    
                    del hash_indexes[col][col_value]
          for col in btree_indexes:
            col_value = row.get(col)
      
            if col_value in btree_indexes[col]["values"]:
                value_list = btree_indexes[col]["values"][col_value]
            
                for i in range(len(value_list) - 1, -1, -1):
                    if self._rows_equal(value_list[i], row):
                        value_list.pop(i)
                        
                        break
                if not value_list:
                    del btree_indexes[col]["values"][col_value]
                    btree_indexes[col]["keys"].remove(col_value)
                    
        
          deleted_count += 1
    def sorted_row(self,rows,order_cluse):
       def compare_rows(row1,row2):
          for clause in order_cluse:
             col=clause["column"]
             dire=clause["direction"]
             val1=row1.get(col)
             val2=row2.get(col)
             if val1 is None and val2 is None:
                continue
             if val1 is None:
                 return 1
             if val2 is None:
                return -1
             try:
                val1=int(val1)
                val2=int(val2)
             except:
                pass
             if val1<val2:
                result= -1
             elif val1>val2:
                result= 1
             else:
                continue
             
             if dire=="DESC":
               result= -result
             return result
          return 0
       sorted_rows=sorted(rows,key=cmp_to_key(compare_rows))
       return sorted_rows
    
    def distinc_on(self,parsed):
       table=parsed["table"]
       distinct=parsed["distinct"]
       select_col=parsed["select_col"]
       order_claues=parsed["order_by"]
       table_file = f"{self.db_path}/{table}.json"
    
       if not os.path.exists(table_file):
        print(f"❌ {table} does not exist")
        return
       all_rows=self.get_allrows(table)
       if order_claues:
          all_rows=self.sorted_row(all_rows,order_claues)
       print(all_rows)
       seen_combination=set()
       result=[]
       for row in all_rows:
          combination=tuple(row[col] for col in distinct)
          if combination not in seen_combination:
             seen_combination.add(combination)
             if select_col=="*":
               result.append(row)
             else:
              fillter_row={}
              for col in select_col:
                if col in row:
                 fillter_row[col]=row[col]
              result.append(fillter_row)
       self._print_dict_table(table,result)
       return result

    def group_by(self,table_name,group_col,aggregates,where_filter=None):
        table_file = f"{self.db_path}/{table_name}.json"
        if not os.path.exists(table_file):
          print(f"❌ Table '{table_name}' does not exist")
          return
    
        with open(table_file, "r") as f:
         table_data = json.load(f)
    
        columns = table_data["columns"]
        rows = table_data["rows"]
        if len(columns)==0 or len(rows)==0:
          print("❌ Columns or Rows are empty")
          return 
        columns_name=list(columns.keys())
        all_rows=self.get_allrows(table_name,group_col)
        if all_rows==None:
         return
        if group_col not in columns_name:
           print(f"❌ column {group_col} does not exist")
           return
        if where_filter:
           col = where_filter["col"]
           op = where_filter["op"]
           value = where_filter["value"]
           filter_rows=[]
           for row in all_rows:
              if op=="=" and str(row.get(col))==str(value):
                filter_rows.append(row)
              elif op==">" and float(row.get(col))>float(value):
                filter_rows.append(row)
              elif op=="<" and float(row.get(col))<float(value):
                 filter_rows.append(row)
              elif op==">=" and float(row.get(col))>=float(value):
                filter_rows.append(row)
              elif op=="<=" and float(row.get(col))<=float(value):
                filter_rows.append(row)
              elif op=="!=" and str(row.get(col))!=str(value):
                filter_rows.append(row)
             
           all_rows=filter_rows
           if not all_rows:
            print(f"❌ No rows match WHERE {col} {op} {value}")
            return []
        groups={}
        for row in all_rows:
           key=str(row[group_col])
           if key not in groups:
               groups[key]=[]      
           groups[key].append(row) 
        results=[]
        for key,group_rows in groups.items():
           result_row={group_col:key}
           for agg in aggregates:
            func = agg["function"]
            col = agg["column"]
            alias = agg["alias"]   
            if col=="*":
               result_row[alias]=len(group_rows)
            else:
               values=[]
               for row in group_rows:
                  if row.get(col) is not None:
                     values.append(row[col])
                  else:
                     print(f"❌ column {col} does not exist in some rows")
                     return
               if func.upper()=="COUNT":
                     result_row[alias]=len(values)
               elif func.upper()=="SUM":
                     result_row[alias]=sum(float(v)for v in values)
               elif func.upper()=="AVG":
                     if values:
                      avg=sum(float(v)for v in values)/len(values)
                      result_row[alias]=avg
                     else:
                        result_row[alias]=0
               elif func.upper()=="MAX":
                     if values:
                        max_val=max(float(v) for v in values)
                        result_row[alias]=max_val
                     else:
                        result_row[alias]=None
               elif func.upper()=="MIN":
                     if values:
                        min_val=min(float(v) for v in values)
                        result_row[alias]=min_val
                     else:
                        result_row[alias]=None
               else:
                   print(f"❌ unsupported aggregate function: {func}")
                   return
           results.append(result_row)  
        print(results)     
        self._print_dict_table(table_name,results)
        return results
               
    def filter_rows(self, table_name,col,op,value):

       table_file = f"{self.db_path}/{table_name}.json"
       with open(table_file, "r", encoding="utf-8") as f:
         table_data = json.load(f)
    
       columns = table_data["columns"]
       col_type = columns.get(col)
       if col_type=="INT":
          con_val=int(value)
       elif col_type=="FLOAT":
          con_val=float(value)
       else:
          con_val=str(value)
      
       if table_name not in self.memory_indexes:
          self._load_indexes(table_name)
       results=[]
       if op =="=" and col in self.memory_indexes[table_name]["hash"]:
          result=self.memory_indexes[table_name]["hash"][col].get(con_val)
          if result:
             results=[result]
          else:
             results=[]
       elif op in [">", ">=", "<", "<=","="] and col in self.memory_indexes[table_name]["b_tree"]:
          btree = self.memory_indexes[table_name]["b_tree"][col]
          if op == ">":
            start_idx = bisect.bisect_right(btree["keys"], con_val)
            for i in range(start_idx, len(btree["keys"])):
                key = btree["keys"][i]
                results.extend(btree["values"][key])
            
          elif op =="=":
              start_idx = bisect.bisect_left(btree["keys"], con_val)
              for i in range(start_idx, len(btree["keys"])):
                if con_val==btree["keys"][i]:
                  results.extend(btree["values"][con_val])
        
          elif op == ">=":
            start_idx = bisect.bisect_left(btree["keys"], con_val)
            for i in range(start_idx, len(btree["keys"])):
                key = btree["keys"][i]
                results.extend(btree["values"][key])
        
          elif op == "<":
            end_idx = bisect.bisect_left(btree["keys"], con_val)
            for i in range(0, end_idx):
                key = btree["keys"][i]
                results.extend(btree["values"][key])
          elif op == "<=":
            end_idx = bisect.bisect_right(btree["keys"], con_val)
            for i in range(0, end_idx):
                key = btree["keys"][i]
                results.extend(btree["values"][key])
       elif op == "!=" and col in self.memory_indexes[table_name]["hash"]:
        for key, row in self.memory_indexes[table_name]["hash"][col].items():
            if key != con_val:
                results.append(row)
      
       else:
          all_rows=self.get_allrows(table_name)
          for row in all_rows:
             if op=="=" and str(row.get(col))==str(value):
                results.append(row)
             elif op==">" and float(row.get(col))>float(value):
                results.append(row)
             elif op=="<" and float(row.get(col))<float(value):
                results.append(row)
             elif op==">=" and float(row.get(col))>=float(value):
                results.append(row)
             elif op=="<=" and float(row.get(col))<=float(value):
                results.append(row)
             elif op=="!=" and str(row.get(col))!=str(value):
                results.append(row)
       if len(results)==0:
          print("no matching rows found ❌")
       else:
         print(f"✅ found {len(results)} matching rows:")
       
       #self._print_dict_table(table_name, results)
       return results
    

    

    def limit(self, table_name,limit=None,offset=None):
     table_file = f"{self.db_path}/{table_name}.json"
     if not os.path.exists(table_file):
        #print(f"❌ Table '{table_name}' does not exist")
        return[]
    
     with open(table_file, "r") as f:
       table_data = json.load(f)
    
     columns = table_data["columns"]
     rows = table_data["rows"]
     if len(columns)==0 or len(rows)==0:
        #print("❌ Columns or Rows are empty")
        return []
     sorted_rows = sorted(rows, key=lambda row:row[0])
     if limit is None:
        result=self._print_table(table_name, sorted_rows)
     else:
        result=self._print_table(table_name, sorted_rows,limit,offset)
     return result if result else []


    def order_by_withcol(self, table_name, order_col,limit=None,offset=None):
     table_file = f"{self.db_path}/{table_name}.json"
     if not os.path.exists(table_file):
        print(f"❌ Table '{table_name}' does not exist")
        return
    
     with open(table_file, "r") as f:
       table_data = json.load(f)
    
     columns = table_data["columns"]
     rows = table_data["rows"]
     if len(columns)==0 or len(rows)==0:
        print("❌ Columns or Rows are empty")
        return False
     if order_col not in columns:
         print(f"column {order_col} does not exist")
         return
     col_names = list(columns.keys())
     col_index = col_names.index(order_col)

     sorted_rows = sorted(rows, key=lambda row: row[col_index])
     if limit is None:
        result=self._print_table(table_name, sorted_rows)
     else:
        result=self._print_table(table_name, sorted_rows,limit,offset)
        
     return result if result else []

    def order_by_full(self, table_name, order_col, order_dir="ASC",limit=None,offset=None):
     table_file = f"{self.db_path}/{table_name}.json"
     if not os.path.exists(table_file):
        print(f"❌ Table '{table_name}' does not exist")
        return
    
     with open(table_file, "r") as f:
       table_data = json.load(f)
    
     columns = table_data["columns"]
     rows = table_data["rows"]
     if len(columns)==0 or len(rows)==0:
        print("❌ Columns or Rows are empty")
        return False
     if order_col not in columns:
         print(f"column {order_col} does not exist")
         return
    
     col_names = list(columns.keys())
     col_index = col_names.index(order_col)

     sorted_rows = sorted(rows, key=lambda row: row[col_index], reverse=(order_dir == "DESC"))
     if limit is None:
        result=self._print_table(table_name, sorted_rows)
     else:
        result=self._print_table(table_name, sorted_rows,limit,offset)
     
     return result if result else []
    



    def _print_dict_table(self, table_name, dict_rows):

     if not dict_rows:
        print("📋 No rows to display")
        return
    
     column_names = list(dict_rows[0].keys())
    
     widths = {}
     for col in column_names:
        widths[col] = len(col)  
        for row in dict_rows:
            value_len = len(str(row.get(col, "")))
            if value_len > widths[col]:
                widths[col] = value_len
    

     total_width = sum(widths.values()) + (len(column_names) * 3) + 1
    
     #print("\n" + "=" * total_width)
     table_title = f"📊 Table: {table_name.upper()}"
     padding = (total_width - len(table_title)) // 2
     #print(" " * padding + table_title)
     #print("=" * total_width)
    

     border = "+"
     for col in column_names:
        border += "-" * (widths[col] + 2) + "+"
    
     #print(border)
    
     header = "|"
     for col in column_names:
        header += f" {col:<{widths[col]}} |"
     #print(header)
    
     #print(border)
    

     for row in dict_rows:
        row_str = "|"
        for col in column_names:
            value = str(row.get(col, ""))
            row_str += f" {value:<{widths[col]}} |"
        print(row_str)
    
     #print(border)
     #print(f"Total rows: {len(dict_rows)}\n")
     return dict_rows


    def _print_table(self, table_name,sorted_rows=None, limit=None, offset=None):
     
     table_file = f"{self.db_path}/{table_name}.json"

     if not os.path.exists(table_file):
        print(f"❌ Table '{table_name}' does not exist")
        return []
    
    
     with open(table_file, "r", encoding="utf-8") as f:
        table_data = json.load(f)

     if sorted_rows==None and limit==None and offset==None : 
      columns = table_data.get("columns", {})
      rows = table_data.get("rows", [])
      if len(columns)==0 or len(rows)==0:
        print("❌ Columns or Rows are empty")
        return[]
      rows=sorted(rows,key=lambda row: row[0])
      column_names = list(columns.keys())
      all_rows = [dict(zip(column_names, row)) for row in rows]
      self._print_formatted_table(table_name, all_rows, column_names)
      return all_rows
   

     elif sorted_rows!=None and limit==None and offset==None:
        columns = table_data.get("columns", {})
        rows = sorted_rows
        column_names = list(columns.keys())
        all_rows = [dict(zip(column_names, row)) for row in rows]
        self._print_formatted_table(table_name, all_rows, column_names)
        return all_rows
    
     elif sorted_rows!=None and limit!=None or offset!=None:
      columns = table_data.get("columns", {})
      rows = sorted_rows
      rows=rows[0:int(limit)] if offset==None else rows[int(limit):int(offset)+1]
      column_names = list(columns.keys())
      all_rows = [dict(zip(column_names, row)) for row in rows]
      self._print_formatted_table(table_name, all_rows, column_names)
     return all_rows
        
    def _print_formatted_table(self, table_name, all_rows, column_names):
       if not all_rows:
        return
        
       widths = {}
       for col in column_names:
        widths[col] = len(col)
        for row in all_rows:
            value_len = len(str(row.get(col, "")))
            if value_len > widths[col]:
                widths[col] = value_len
    
       total_width = sum(widths.values()) + (len(column_names) * 3) + 1
    
       #print("\n" + "=" * total_width)
       table_title = f"📊 Table: {table_name.upper()}"
       padding = (total_width - len(table_title)) // 2
       #print(" " * padding + table_title)
       #print("=" * total_width)
    
       border = "+"
       for col in column_names:
        border += "-" * (widths[col] + 2) + "+"
       #print(border)
    
       header = "|"
       for col in column_names:
        header += f" {col:<{widths[col]}} |"
       #print(header)
       #print(border)
    
       for row in all_rows:
        row_str = "|"
        for col in column_names:
            value = str(row.get(col, ""))
            row_str += f" {value:<{widths[col]}} |"
        #print(row_str)
    
       #print(border)
       print(f"Total rows: {len(all_rows)}\n")
       
      
